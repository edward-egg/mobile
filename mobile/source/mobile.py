import csv
import re
import time
from collections import Counter
import wx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

wb = Workbook()
r = r'(201801).+'

yellow_fill = PatternFill(fill_type='solid',fgColor='FFFF00')
green_fill = PatternFill(fill_type='solid',fgColor='94D050')

def init_excel(csvFile, date = 201801,time1 = time.strftime('%m%d', time.localtime())):
    wb = Workbook()
    sheet = wb.create_sheet('社区通报'+time1, 0)

    #选择日期
    r =r'(' + str(date) + ').+'

    #引入类型的id
    type = init_type('type.csv')

    #记录
    count = 0
    #读取csv
    reader = csv.reader(open(csvFile))

    #插入宽带类型和社区经理列
    for i,item in enumerate(reader):
        if i == 0:
            item.insert(26,'宽带类型')
            item.insert(63,'社区经理')
            sheet.append(item)
        if(re.search(r,item[25]) and re.search(r,item[36])):
            item.insert(26, '')
            item.insert(63, '')
            item[22:26] = item[34:38]
            item[27:29] = item[38:40]
            item[30:34] = item[40:44]
            sheet.append(item)
            count += 1


    t = 1
    #用当前方案id进行宽带类型匹配
    for cell in list(sheet.columns)[22]:
        if cell.value in type[0]:
            sheet['AA'+ str(t)] = '极光新装'
            t +=1
        elif cell.value in type[1]:
            sheet['AA' + str(t)] = '极光续费'
            t += 1
        elif cell.value in type[2]:
            sheet['AA' + str(t)] = '普通新装'
            t += 1
        elif cell.value in type[3]:
            sheet['AA' + str(t)] = '普通续费'
            t += 1
        else:
            sheet['AA' + str(1)] = '宽带类型'
            t += 1

    wb.save('社区通报.xlsx')


#引入宽带类型
def init_type(csvFile):
    # type_list = []
    # reader = csv.reader(open(csvFile))
    #
    # for i,item in enumerate(reader):
    #     if i == 0:
    #         continue
    #     type_list.append(item[0])
    # return type_list
    type_list = []
    column = []
    j = 0
    csv_file = open(csvFile)
    t = len(csv_file.readline().split(','))
    for item in range(0, t, 3):
        col = []
        csv_file = open(csvFile)
        reader = csv.reader(csv_file)
        csv_file.readline()
        for i,item in enumerate(reader):
            if item[j] != '':
                col.append(item[j])
            else:
                break;
        type_list.append(col)
        if j < t:
            j += 3
        else:
            break;
    return type_list


#引入社区经理
def init_manager(csvFile):

    # 客户经理数组
    name = []
    column = []
    #控制列号
    j = 0
    csv_file = open(csvFile)
    #最大列号
    t = len(csv_file.readline().split(','))
    for item in range(0,t,3):
        col = []
        csv_file = open(csvFile)
        reader = csv.reader(csv_file)
        csv_file.readline()
        for i,item in enumerate(reader):
            if i==0:
                name.append(item[j+1])
            if item[j] != '':
                col.append(item[j])
            else:
                break;
        column.append(col)
        if j<t:
            j += 3
        else:
            break;
    return column,name

def fill(workbook='社区通报.xlsx',save1='汇总报表.xlsx',outfit = 9,header_col = 13,time1 = time.strftime('%m%d', time.localtime())):

    wb =load_workbook(workbook)
    sheet = wb['社区通报'+time1]
    manager1,name1 = init_manager('manager_1.csv')
    manager2,name2 = init_manager('manager_2.csv')
    manager1.extend(manager2)
    name1.extend(name2)
    list1 = []
    for i in range(0,len(name1)):
        list1.append([])
    words = [chr(i) for i in range(65, 91)]

    #操作地址id列
    for j,item in enumerate(list(sheet.columns)[57]):
        for i,v in enumerate(manager1):
            if item.value in v:
                sheet['BL'+str(j+1)] = name1[i]
                list1[i].append(sheet['AA'+str(j+1)].value)
                break

    #记录内容分列总数
    sum1 = sum2 = sum3 = sum4 = 0
    sheet2 = wb.create_sheet('汇总报表', 0)

    #表头
    sheet2.cell(row=outfit,column=header_col+1).value = "普通新装"
    sheet2.cell(row=outfit, column=header_col+2).value ="普通续费"
    sheet2.cell(row=outfit, column=header_col+3).value ="极光新装"
    sheet2.cell(row=outfit, column=header_col+4).value = "极光续费"
    sheet2.cell(row=outfit, column=header_col+5).value =  '总计'
    sheet2.cell(row=outfit+25, column=header_col).value =  '总计'


    for i,item in enumerate(sheet2[ words[header_col-1] + str(outfit+1) : words[header_col+4] + str(outfit+24)]):
        c = Counter(list1[i])
        item[0].value = name1[i]
        item[1].value = c["普通新装"]
        item[2].value = c["普通续费"]
        item[3].value = c["极光新装"]
        item[4].value = c["极光续费"]
        item[5].value = item[1].value+item[2].value+item[3].value+item[4].value
        sum1 +=item[1].value
        sum2 +=item[2].value
        sum3 +=item[3].value
        sum4 +=item[4].value
    sheet2[words[header_col] + str(outfit+25)] = sum1
    sheet2[words[header_col+1] + str(outfit+25)] = sum2
    sheet2[words[header_col+2] + str(outfit+25)] = sum3
    sheet2[words[header_col+3] + str(outfit+25)] = sum4
    wb.save(save1)


class StaticTextFrame(wx.Frame):
    global content
    def __init__(self):

        wx.Frame.__init__(self, None, -1, u'Excel 自动化', size=(350, 500))
        box_sizer = wx.WrapSizer()
        self.SetAutoLayout(True)
        self.SetSizer(box_sizer)


        static_text = wx.StaticText(self, -1, u'文件路径', style=wx.ALIGN_CENTER)
        static_text.SetForegroundColour('white')  # 颜色
        wx_font = wx.Font(16,wx.DEFAULT,wx.NORMAL,wx.BOLD)
        static_text.SetFont(wx_font)
        box_sizer.Add(static_text)


        input_text = wx.TextCtrl(self, -1, size=(300, -1))
        input_text.SetInsertionPoint(0)
        self.Bind(wx.EVT_TEXT,self.onKeyType,input_text)
        box_sizer.Add(input_text)


        #
        # static_text = wx.StaticText(self, -1, u'客户经理', style=wx.ALIGN_CENTER)
        # static_text.SetForegroundColour('white')  # 颜色
        # wx_font = wx.Font(16, wx.DEFAULT, wx.NORMAL, wx.BOLD)
        # static_text.SetFont(wx_font)
        # box_sizer.Add(static_text)
        #
        # input_text = wx.TextCtrl(self, -1, u'', size=(300, -1))
        # input_text.SetInsertionPoint(0)
        # self.content_manager1 = input_text.GetValue()
        # box_sizer.Add(input_text)
        #
        # input_text = wx.TextCtrl(self, -1, u'', size=(300, -1))
        # input_text.SetInsertionPoint(0)
        # self.content_manager2 = input_text.GetValue()
        # box_sizer.Add(input_text)
        #
        # static_text = wx.StaticText(self, -1, u'普通续费', style=wx.ALIGN_CENTER)
        # static_text.SetForegroundColour('white')  # 颜色
        # wx_font = wx.Font(16, wx.DEFAULT, wx.NORMAL, wx.BOLD)
        # static_text.SetFont(wx_font)
        # box_sizer.Add(static_text)
        #
        # input_text = wx.TextCtrl(self, -1, u'', size=(300, -1))
        # input_text.SetInsertionPoint(0)
        # self.content_kind1 = input_text.GetValue()
        # box_sizer.Add(input_text)
        #
        # static_text = wx.StaticText(self, -1, u'普通新装', style=wx.ALIGN_CENTER)
        # static_text.SetForegroundColour('white')  # 颜色
        # wx_font = wx.Font(16, wx.DEFAULT, wx.NORMAL, wx.BOLD)
        # static_text.SetFont(wx_font)
        # box_sizer.Add(static_text)
        #
        # input_text = wx.TextCtrl(self, -1, u'', size=(300, -1))
        # input_text.SetInsertionPoint(0)
        # self.content_kind2 = input_text.GetValue()
        # box_sizer.Add(input_text)
        #
        # static_text = wx.StaticText(self, -1, u'极光新装', style=wx.ALIGN_CENTER)
        # static_text.SetForegroundColour('white')  # 颜色
        # wx_font = wx.Font(16, wx.DEFAULT, wx.NORMAL, wx.BOLD)
        # static_text.SetFont(wx_font)
        # box_sizer.Add(static_text)
        #
        # input_text = wx.TextCtrl(self, -1, u'', size=(300, -1))
        # input_text.SetInsertionPoint(0)
        # self.content_kind3 = input_text.GetValue()
        # box_sizer.Add(input_text)
        #
        # static_text = wx.StaticText(self, -1, u'极光续费', style=wx.ALIGN_CENTER)
        # static_text.SetForegroundColour('white')  # 颜色
        # wx_font = wx.Font(16, wx.DEFAULT, wx.NORMAL, wx.BOLD)
        # static_text.SetFont(wx_font)
        # box_sizer.Add(static_text)
        #
        # input_text = wx.TextCtrl(self, -1, u'', size=(300, -1))
        # input_text.SetInsertionPoint(0)
        # self.content_kind4 = input_text.GetValue()
        # box_sizer.Add(input_text)


        # button1 = wx.Button(self,-1,u'Choose',pos=(10,20),size=(80,30))
        # self.Bind(wx.EVT_BUTTON,self.OnClick1,button1)
        # box_sizer.Add(button1)

        button2 = wx.Button(self,-1,u'Submit',pos=(30,50),size=(80,30))
        self.Bind(wx.EVT_BUTTON,self.OnClick2,button2)
        box_sizer.Add(button2)


    def OnClick2(self,event):
        init_excel(self.content)
        fill()

    # def OnClick1(self,event):
    #     self.SetLabel("clicked")

    def onKeyType(self,event):
        self.content = event.GetString()
        print(self.content)
        return self.content



if __name__ == '__main__':
    root = wx.App()
    frame = StaticTextFrame()
    frame.Show()
    root.MainLoop()




